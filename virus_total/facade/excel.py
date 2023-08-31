from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from .base import Facade
from datetime import datetime
from os.path import join

class ExcelFacade(Facade):
  def open(self):
    return super().open('Abrir hoja de calculo', (('Documento de Excel (2007-posterior)', '*.xlsx'), ('Documento de Excel 2003', '*.xls')))


  def load(self, path):
    return load_workbook(path)


  def res_file(self, **kwargs):
    return super().res_file('Guardar hoja de calculo', '.xlsx', **kwargs)


  def save(self, path, **kwargs):
    wb = kwargs.get('data', None)
    wb.save(join(path, datetime.now().strftime('%d%m%Y_%H%M%S.xlsx')))


  def custom_process(self, headers, data, **kwargs):
    def __format_cell__(cell):
      cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
      if 'Error' in str(cell.value):
        cell.font = Font(color='626262')
        cell.fill = PatternFill(fill_type='solid', start_color="CBCBCB", end_color="CBCBCB")
      elif 'Warning' in str(cell.value):
        cell.font = Font(color='AC9A00')
        cell.fill = PatternFill(fill_type='solid', start_color="FFF17C", end_color="FFF17C")
    
    sheetname = kwargs.get('sheetname', 'Hoja')
    wb = kwargs.get('wb', Workbook())
    ws = wb[sheetname] if sheetname in wb.sheetnames else wb.create_sheet(sheetname)

    ws.append(headers)

    for row in data:
      ws.append(row)

    rules = {
      'start_type': 'percentile',
      'start_value': 0,
      'start_color': 'A2FFA2',
      'mid_type': 'percentile',
      'mid_value': 80,
      'mid_color': 'FFA2A2',
      'end_type': 'percentile',
      'end_value': 100,
      'end_color': 'FFA2A2'
    }

    rg = f'C2:C{ws.max_row}'
    ws.conditional_formatting.add(rg, ColorScaleRule(**rules))

    
    for cell in list(ws['C']):
      __format_cell__(cell)
    
    return wb
    

  
