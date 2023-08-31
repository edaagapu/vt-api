from tkinter import Tk, Menu, messagebox as MessageBox
from .components import ApplicationFrame, ImportationFrame
from .views import SettingsView
from os.path import join, dirname, abspath
from encrypt import JWTEncrypt
from api import *
from openpyxl import load_workbook, Workbook
from facade import JSONFacade, ExcelFacade

_ICON_FILE = join(dirname(abspath(__file__)), 'icons\main.ico')

class AppController:
  def set_path(self, path):
    self.app_frame.path_text.set(path)

  def select_provider(self, provider):
    if provider == 'IBM':
      return IBMAPI()
    elif provider == 'VT':
      return VTAPI()
    
  def import_data(self):
    wb = self.facade.load(self.app_frame.path_text.get())
    ws = wb.active
    return list(map(lambda cell: cell.value, ws['A']))
  
  def classify_to_dict(self, values, is_classified=False):
    if is_classified:
      r_results = values
    else:
      r_results = self.classify(values)
    
    results = {
      'dns': [],
      'hash': [],
      'ip': [],
      'unknown': [],
      'url': [],
    }

    for value, value_type, is_repeat in r_results:
      if not is_repeat:
        if 'hash' in value_type:
          results['hash'].append(value)
        else:
          results[value_type].append(value)
    return results


  def classify(self, values):
    results = []
    for index in range(len(values)):
      category = 'unknown'
      if VTAPI.__isipv4__(values[index]):
        category = 'ip'
      elif VTAPI.__ishash__(values[index]):
        category = f'hash - {VTAPI.hash_type(values[index])}'
      elif VTAPI.__isdns__(values[index]):
        category = 'dns'
      elif VTAPI.__isurl__(values[index]):
        category = 'url'
      results.append((
        values[index],
        category,
        not (values.count(values[index]) == 1 or values.index(values[index]) == index)
      ))
    return results

  def results_hash(self, r_values, provider, provider_key):
    results = []
    values = [[r_value, False] for r_value in r_values]
    for index in range(len(values)):
      if not values[index][1]:
        try:
          data_resp = provider.consult_result(values[index][0], provider_key).json()
          hash_types = ['sha1', 'sha256', 'md5']
          original_type = provider.hash_type(values[index][0])
          hash_types.remove(original_type)
          for ht in hash_types:
            o_hash = data_resp.get('data').get('attributes').get(ht)
            if o_hash in r_values:
              values[r_values.index(o_hash)][1] = True
          
          res = data_resp.get('data').get('attributes')
          label = None
          categories = None
          if 'popular_threat_classification' in res.keys():
            label = res.get('popular_threat_classification').get('suggested_threat_label')
            categories = ','.join([category.get('value') for category in res.get('popular_threat_classification').get('popular_threat_category')])

          malicious_v = res.get('last_analysis_stats').get('malicious')
          total_v = (malicious_v + res.get('last_analysis_stats').get('undetected'))
          ranking = (malicious_v / total_v) * 100 if total_v != 0 else 0
          results.append({
            'valor': values[index][0],
            'tipo': f'hash - {original_type}',
            'ranking': ranking,
            'hash_md5': res.get('md5'),
            'hash_sha1': res.get('sha1'),
            'hash_sha256': res.get('sha256'),
            'etiqueta': label,
            'categorias': categories,
            'comentario': ''
          })
        except Exception as error:
          results.append({
            'valor': values[index][0],
            'tipo': None,
            'ranking': ranking,
            'hash_md5': None,
            'hash_sha1': None,
            'hash_sha256': None,
            'etiqueta': None,
            'categorias': None,
            'comentario': str(error)
          })
      else:
        results.append({
          'valor': values[index][0],
          'tipo': None,
          'hash_md5': None,
          'hash_sha1': None,
          'hash_sha256': None,
          'etiqueta': None,
          'categorias': None,
          'ranking': None,
          'comentario': 'Error: Valor duplicado'
        })
    
    return results

  def results_ip(self, ip_values, provider, provider_keys):
    pass

  def run_api(self):
    encrypt = JWTEncrypt()
    key = encrypt.load_key()
    settings = encrypt.load_settings(key)
    
    providers = list(map(lambda x: x.split('_')[0].upper(),filter(lambda x: 'key' in x and not 'vt' in x, settings.keys() )))
    # Fist should be VT
    providers.insert(0, 'VT')
    
    filepath = settings.get('export_path')

    wb = Workbook()
    ws = wb.get_sheet_by_name('Sheet')
    wb.remove(ws)
    values = self.import_data()
    
    cr_values = self.classify(values)
    ws = wb.create_sheet('Original')
    for value in cr_values:
      ws.append(value)

    c_values = self.classify_to_dict(cr_values, True)
    hash_values = c_values.pop('hash', None)
    
    for provider in providers:
      provider_keys = settings.get(f'{provider.lower()}_keys')
      api = self.select_provider(provider)
      dcy_pk = encrypt.decrypt(key=key, token=provider_keys[0])
      data = self.facade.process_data(self.results_hash(hash_values, api, dcy_pk), sheetname=provider, wb=wb)
      self.facade.save(filepath, data=data)
 
class App(Tk, AppController):
  in_work = False

  def __init__(self, facade=ExcelFacade(), *args, **kwargs):
    super().__init__(*args, **kwargs)
    self.facade = facade
    self.title('Cerberus')
    self.create_widgets()
    self.create_menubar()

    self.config(menu=self.menubar, width=500, height=500)
    self.iconbitmap(_ICON_FILE)

  def create_menubar(self):
    menu_items = {
      'Archivo': [('Importar archivo', self.imp_frame.import_file), '/', ('Salir', self.destroy)],
      'Configuración': [('Propiedades', self.run_api)],
      'Ayuda': ['Ayuda', '/', 'Acerca de']
    }

    self.menubar = Menu(self)

    for item in menu_items.keys():
      submenu = Menu(self.menubar, tearoff=0)
      for subitem in menu_items[item]:
        if type(subitem) == str:
          if subitem == '/':
            submenu.add_separator()
          else:
            submenu.add_command(label=subitem)
        elif type(subitem) == tuple:
          submenu.add_command(label=subitem[0], command=subitem[1])
      self.menubar.add_cascade(label=item, menu=submenu)

  def create_widgets(self):
    feature = {'bg': '#E9E9E9', 'bd': 1, 'relief': 'solid'}
    self.app_frame = ApplicationFrame(self, color=feature.get('bg'))
    self.imp_frame = ImportationFrame(self, color=feature.get('bg'), facade=self.facade)
    self.app_frame.config(**feature)
    self.imp_frame.config(**feature)

  def destroy(self):
    result = True
    while (SettingsView.in_use):
      self.imp_frame.destroy()
    if self.__class__.in_work:
      result = MessageBox.askokcancel('Salir', '¿Desea salir de la aplicación sin terminar de ejecutar el proceso?')
    if result:
      super().destroy()
